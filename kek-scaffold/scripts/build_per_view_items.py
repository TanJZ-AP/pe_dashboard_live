"""Build per-view item rankings: DATA.top20_by_view and DATA.top10_wib_by_view.

The Top 20 KEK + Top 10 WIB tables are view-aware (filter by MTD/W1/W2/W3/W4).
This script aggregates items from ALL sources (per-day menuItemReports + DayPartGroup
files + OrderType + WIB DayPartReport) into items_by_date, then sums per view.

Run after pull_latest.py whenever new menuItemReport files arrive.
"""
import openpyxl, os, re, json

KEK_DIR = "/sessions/youthful-peaceful-faraday/mnt/Daily POS Reports"
SCRIPTS = os.path.dirname(os.path.abspath(__file__))
UPLOADS = "/sessions/youthful-peaceful-faraday/mnt/uploads"
OUTLETS = ['KEK Alexandra', 'KEK Tampines', 'KEK Punggol', 'WOKIN Punggol']
SHORT = {'KEK Alexandra': 'BM', 'KEK Tampines': 'TP', 'KEK Punggol': 'PG', 'WOKIN Punggol': 'WIB'}

W1_DATES = ["2026-03-30","2026-03-31","2026-04-01","2026-04-02","2026-04-03","2026-04-04","2026-04-05"]
W2_DATES = ["2026-04-06","2026-04-07","2026-04-08","2026-04-09","2026-04-10","2026-04-11","2026-04-12"]
W3_DATES = ["2026-04-13","2026-04-14","2026-04-15","2026-04-16","2026-04-17","2026-04-18","2026-04-19"]
W4_DATES = ["2026-04-20","2026-04-21","2026-04-22","2026-04-23","2026-04-24","2026-04-25","2026-04-26"]
W5_DATES = ["2026-04-27","2026-04-28","2026-04-29","2026-04-30","2026-05-01","2026-05-02","2026-05-03"]

def num(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def identify(f):
    for k, v in [('Alexandra', 'KEK Alexandra'), ('Punggol SAFRA', 'KEK Punggol'),
                 ('Tampines Safra KEK', 'KEK Tampines'), ('Wok In Burger', 'WOKIN Punggol')]:
        if k in f: return v
    return None

# Build items_by_date[date][outlet][item] = {qty, net}
items_by_date = {}
def add(date, outlet, item, qty, net):
    if not item or net <= 0: return
    items_by_date.setdefault(date, {}).setdefault(outlet, {}).setdefault(item, {'qty': 0, 'net': 0})
    items_by_date[date][outlet][item]['qty'] += qty
    items_by_date[date][outlet][item]['net'] += net

# 1. Per-day menuItemReports
for f2 in sorted(os.listdir(KEK_DIR)):
    if 'menuItemReport' not in f2: continue
    outlet = identify(f2)
    if not outlet: continue
    m = re.search(r'(\d{4}-\d{2}-\d{2})', f2)
    if not m: continue
    date = m.group(1)
    try:
        wb_m = openpyxl.load_workbook(os.path.join(KEK_DIR, f2), data_only=True)
        ws_m = wb_m.active
        cur = None
        for row in ws_m.iter_rows(values_only=True):
            if row[0] == 'Dine In': cur = 'DI'; continue
            if row[0] == 'Take Away': cur = 'TA'; continue
            if not cur: continue
            if row[0] is None and row[1] is not None and row[2] is not None:
                name = str(row[2]).strip()
                try:
                    qty = float(row[3]) if row[3] else 0
                    net = float(row[7]) if len(row) > 7 and row[7] is not None else 0
                except: qty = net = 0
                add(date, outlet, name, qty, net)
    except: pass

# 2. BM/TP/PG DayPartGroup per-day files (Mar 30 - Apr 15)
DATE_MAP = [
    ('2026-03-30','300326'),('2026-03-31','310326'),
    ('2026-04-01','010426'),('2026-04-02','020426'),('2026-04-03','030426'),
    ('2026-04-04','040426'),('2026-04-05','050426'),
    ('2026-04-06','060426'),('2026-04-07','070426'),('2026-04-08','080426'),
    ('2026-04-09','090426'),('2026-04-10','100426'),('2026-04-11','110426'),('2026-04-12','120426'),
    ('2026-04-13','130426'),('2026-04-14','140426'),('2026-04-15','150426'),
]

# BM Mar 30 - Apr 15
for date, mmdd in DATE_MAP:
    fp = f"{UPLOADS}/DayPartGroup_{mmdd}~{mmdd}_Keng Eng Kee Seafood @ Alexandra Village.xlsx"
    if not os.path.exists(fp): continue
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=6, values_only=True):
        day_part, category, pax, item_name, qty, gross, disc, sc, tax, net = row[:10]
        if day_part is None and category is None and item_name and isinstance(item_name, str):
            add(date, 'KEK Alexandra', item_name.strip(), num(qty), num(net))

# TP Mar 30 - Apr 15
for date, mmdd in DATE_MAP:
    for f in os.listdir(UPLOADS):
        if f.startswith(f"DayPartGroup_{mmdd}~{mmdd}_KEK2"):
            wb = openpyxl.load_workbook(os.path.join(UPLOADS, f), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=6, values_only=True):
                day_part, category, pax, item_name, qty, gross, disc, sc, tax, net = row[:10]
                if day_part is None and category is None and item_name and isinstance(item_name, str):
                    add(date, 'KEK Tampines', item_name.strip(), num(qty), num(net))
            break

# PG Mar 30 - Apr 4 (W1, W2 Apr 6+ comes from individual menuItemReports)
for date, mmdd in DATE_MAP[:6]:
    fp = f"{UPLOADS}/DayPartGroup_{mmdd}~{mmdd}_Keng Eng Kee Seafood @ Punggol SAFRA.xlsx"
    if not os.path.exists(fp): continue
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=6, values_only=True):
        day_part, category, pax, item_name, qty, gross, disc, sc, tax, net = row[:10]
        if day_part is None and category is None and item_name and isinstance(item_name, str):
            add(date, 'KEK Punggol', item_name.strip(), num(qty), num(net))

# 3. PG Apr 13 derived from OrderType (OrderType 13-15 minus existing 14+15)
def parse_ot(fpath):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    out = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] in ('DINE IN', 'TAKEAWAY', 'Total') or row[0] != 'undefined': continue
        name = row[2]
        if not name or not isinstance(name, str): continue
        name = name.strip()
        try:
            qty = float(row[3]) if row[3] else 0
            net = float(row[11]) if len(row) > 11 and row[11] is not None else 0
        except: qty = net = 0
        if net <= 0: continue
        if name not in out: out[name] = {'qty': 0, 'net': 0}
        out[name]['qty'] += qty; out[name]['net'] += net
    return out

OT_PG = f"{UPLOADS}/OrderTypeReport_13-4-2026~14-4-2026_Keng Eng Kee Seafood @ Punggol SAFRA.xlsx"
if os.path.exists(OT_PG):
    pg_14 = items_by_date.get('2026-04-14', {}).get('KEK Punggol', {})
    pg_15 = items_by_date.get('2026-04-15', {}).get('KEK Punggol', {})
    for item, ot_v in parse_ot(OT_PG).items():
        k14 = pg_14.get(item, {'qty': 0, 'net': 0})
        k15 = pg_15.get(item, {'qty': 0, 'net': 0})
        n13 = ot_v['net'] - k14['net'] - k15['net']
        q13 = ot_v['qty'] - k14['qty'] - k15['qty']
        if n13 > 0.01:
            add('2026-04-13', 'KEK Punggol', item, q13, n13)

# 4. WIB Apr 13 from DayPartReport
wib_f = f"{UPLOADS}/DayPartReport_13-4-2026~13-4-2026_WOKIN.xlsx"
if os.path.exists(wib_f):
    wb_w = openpyxl.load_workbook(wib_f, data_only=True)
    ws_w = wb_w.active
    for row in ws_w.iter_rows(min_row=7, values_only=True):
        day_part, category, item_name = row[0], row[1], row[2]
        if day_part is not None or category is not None: continue
        if not item_name or not isinstance(item_name, str): continue
        try:
            qty = float(row[3]) if row[3] else 0
            net = float(row[11]) if len(row) > 11 and row[11] is not None else 0
        except: qty = net = 0
        add('2026-04-13', 'WOKIN Punggol', item_name.strip(), qty, net)

# Aggregate per view
def aggregate(dates):
    agg = {}
    for d in dates:
        for outlet, items in items_by_date.get(d, {}).items():
            for name, v in items.items():
                if name not in agg:
                    agg[name] = {o: {'qty': 0, 'net': 0} for o in OUTLETS}
                agg[name][outlet]['qty'] += v['qty']
                agg[name][outlet]['net'] += v['net']
    combined = []
    for item, od in agg.items():
        kek_net = sum(od[o]['net'] for o in ['KEK Alexandra', 'KEK Tampines', 'KEK Punggol'])
        total_net = sum(od[o]['net'] for o in OUTLETS)
        total_qty = sum(od[o]['qty'] for o in OUTLETS)
        combined.append({'item': item, 'net': round(total_net, 2), 'qty': int(total_qty),
                         'kek_net': kek_net, 'outlets': {SHORT[o]: round(od[o]['net'], 2) for o in OUTLETS}})
    combined.sort(key=lambda x: -x['kek_net'])
    top20 = [{'item': x['item'], 'net': x['net'], 'qty': x['qty'], 'outlets': x['outlets']} for x in combined[:20]]
    wib = [(item, agg[item]['WOKIN Punggol']['qty'], agg[item]['WOKIN Punggol']['net'])
           for item in agg if agg[item]['WOKIN Punggol']['net'] > 0]
    wib.sort(key=lambda x: -x[2])
    top10 = [{'item': i, 'qty': int(q), 'net': round(n, 2)} for i, q, n in wib[:10]]
    return top20, top10

all_dates = sorted(items_by_date.keys())
views = {'mtd': all_dates, 'w1': W1_DATES, 'w2': W2_DATES, 'w3': W3_DATES, 'w4': W4_DATES, 'w5': W5_DATES}
top_by_view = {}
top10_by_view = {}
for v, dates in views.items():
    t20, t10 = aggregate(dates)
    top_by_view[v] = t20
    top10_by_view[v] = t10

with open(f'{SCRIPTS}/kek_dashboard_data.json') as f:
    dash = json.load(f)
dash['top20_by_view'] = top_by_view
dash['top10_wib_by_view'] = top10_by_view
dash['top20_kek'] = top_by_view['mtd']
dash['top10_wib'] = top10_by_view['mtd']
with open(f'{SCRIPTS}/kek_dashboard_data.json', 'w') as f:
    json.dump(dash, f, indent=2, ensure_ascii=False)
print(f"Built top items for {len(views)} views from {len(all_dates)} dates")
for v, t in top_by_view.items():
    print(f"  {v}: top1={t[0]['item'][:40]} ${t[0]['net']:,.2f}" if t else f"  {v}: empty")
