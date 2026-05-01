"""Parse KEK POS daily reports into structured data."""
import openpyxl
import os
import re
import json
from collections import defaultdict

KEK_DIR = "/sessions/youthful-peaceful-faraday/mnt/Daily POS Reports"

OUTLETS = {
    'Alexandra': 'KEK Alexandra',
    'Punggol SAFRA': 'KEK Punggol',
    'Tampines Safra KEK': 'KEK Tampines',
    'Wok In Burger': 'WOKIN Punggol',
}


def identify_outlet(fname):
    for k, v in OUTLETS.items():
        if k in fname:
            return v
    return None


def num(v):
    try:
        return float(v) if v is not None and v != '' else 0.0
    except (ValueError, TypeError):
        return 0.0


def parse_daily(fpath):
    """Parse a single dailyReport xlsx and return dict of metrics."""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    # Build a dict: value by label — read row by row
    out = {
        'net_sales': 0.0, 'service_charge': 0.0, 'tax': 0.0, 'total_revenue': 0.0,
        'total_discounts': 0.0,
        'di_net_sales': 0.0, 'ta_net_sales': 0.0,
        'di_covers': 0, 'ta_covers': 0, 'di_checks': 0, 'ta_checks': 0,
        'total_checks': 0,
        'grab_net_sales': 0.0, 'foodpanda_net_sales': 0.0, 'oddle_net_sales': 0.0, 'selfcollect_net_sales': 0.0,
        'lunch_net_sales': 0.0, 'tea_net_sales': 0.0, 'dinner_net_sales': 0.0,
        'lunch_covers': 0, 'tea_covers': 0, 'dinner_covers': 0,
        'tender_credit_card': 0.0,  # VISA/MASTER/AMEX
        'tender_cash': 0.0,
        'tender_digital': 0.0,  # PayLah, Grab Pay, FavePay
        'tender_voucher': 0.0,
        'tender_platform': 0.0,  # GRAB FOOD, FOOD PANDA
        'tender_other': 0.0,
        'menu_categories': {},  # category -> net sales
    }

    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(r)

    def find_row(label_pattern, col=0):
        for r in rows:
            if r[col] and isinstance(r[col], str) and re.search(label_pattern, r[col], re.I):
                return r
        return None

    # Sales Metrics
    r = find_row(r'^Net Sales')
    if r: out['net_sales'] = num(r[2])
    r = find_row(r'^\+?Service Charges\s*$')
    if r: out['service_charge'] = num(r[2])
    r = find_row(r'^\+?Tax Collected')
    if r: out['tax'] = num(r[2])
    r = find_row(r'=?Total Revenue')
    if r: out['total_revenue'] = num(r[2])
    r = find_row(r'^Total Discounts')
    if r: out['total_discounts'] = num(r[2])

    # Checks
    r = find_row(r'\+ Checks Begun')
    if r: out['total_checks'] = int(num(r[1]))

    # Order Type section
    in_order_type = False
    for i, row in enumerate(rows):
        if row[0] == 'Order Type':
            in_order_type = True
            continue
        if in_order_type and row[0] and isinstance(row[0], str):
            if row[0].startswith('Take Away') or row[0] == 'Day Part':
                in_order_type = False
                continue
            label = row[0].split('.')[0].upper().strip()
            net = num(row[1])
            # Parse covers — can be weird concatenated string
            covers_raw = row[3]
            covers = 0
            if isinstance(covers_raw, (int, float)):
                covers = int(covers_raw)
            elif isinstance(covers_raw, str) and covers_raw.isdigit():
                # String of digits — take first 3 chars or so
                # Actually these are guests column; if it's a weird concatenation
                # Take the number from next columns Ave Guest if available
                # For safety parse the sane length
                if len(covers_raw) <= 4:
                    covers = int(covers_raw)
                else:
                    covers = 0  # corrupt
            checks = int(num(row[6]))
            if label == 'DINE IN':
                out['di_net_sales'] = net
                out['di_covers'] = covers
                out['di_checks'] = checks
            elif label == 'TAKEAWAY':
                out['ta_net_sales'] = net
                out['ta_covers'] = covers if covers and covers < 500 else 0
                out['ta_checks'] = checks
            elif label == 'TOTAL':
                # skip — already have net_sales
                pass
            elif label == 'GRAB':
                out['grab_net_sales'] = net
                out['ta_net_sales'] += net  # include platform sales in TA total
            elif label == 'FOODPANDA':
                out['foodpanda_net_sales'] = net
                out['ta_net_sales'] += net
            elif label == 'ODDLE':
                out['oddle_net_sales'] = net
                out['ta_net_sales'] += net
            elif label == 'DELIVEROO':
                out['ta_net_sales'] += net

    # Take Away subcategories (SELF COLLECTION etc.)
    in_ta = False
    for row in rows:
        if row[0] == 'Take Away':
            in_ta = True
            continue
        if in_ta and row[0] and isinstance(row[0], str):
            if row[0] == 'Day Part' or row[0] == 'Tender Type':
                in_ta = False
                continue
            label = row[0].split('.')[0].upper().strip()
            net = num(row[1])
            if label == 'SELF COLLECTION':
                out['selfcollect_net_sales'] = net
            elif label == 'GRAB FOOD':
                out['grab_net_sales'] = max(out['grab_net_sales'], net)
            elif label == 'FOOD PANDA':
                out['foodpanda_net_sales'] = max(out['foodpanda_net_sales'], net)

    # Day Part section
    in_dp = False
    for row in rows:
        if row[0] == 'Day Part':
            in_dp = True
            continue
        if in_dp and row[0] and isinstance(row[0], str):
            if row[0] == 'Tender Type':
                in_dp = False
                continue
            label = row[0].upper().strip()
            net = num(row[1])
            covers = int(num(row[3]))
            if label == 'LUNCH':
                out['lunch_net_sales'] = net
                out['lunch_covers'] = covers
            elif label == 'TEA':
                out['tea_net_sales'] = net
                out['tea_covers'] = covers
            elif label == 'DINNER':
                out['dinner_net_sales'] = net
                out['dinner_covers'] = covers

    # Tender Type section — classify
    in_tender = False
    CC_NAMES = {'VISA', 'MASTER', 'AMEX', 'CC'}
    DIGITAL_NAMES = {'PAYLAH', 'GRAB PAY', 'FAVEPAY', 'NETS', 'PAYNOW', 'CDC'}
    VOUCHER_NAMES = {'SAFRA VOUCHER', 'KEK VOUCHER', 'STB EVENT', 'SAFRA BIRTHDAY VOUCHER', 'VOUCHER', 'STAFF MEAL', 'KEK DRINK', 'SAFRA (INTERNAL)'}
    PLATFORM_NAMES = {'GRAB FOOD', 'FOOD PANDA', 'ODDLE', 'DELIVEROO'}
    for row in rows:
        if row[0] == 'Tender Type':
            in_tender = True
            continue
        if in_tender and row[0] and isinstance(row[0], str):
            if row[0].startswith('Major Group') or row[0].startswith('TOTAL'):
                if row[0].startswith('Major'):
                    in_tender = False
                    continue
                continue  # skip TOTAL row
            label = row[0].split('.')[0].upper().strip()
            amt = num(row[1])
            matched = False
            for cc in CC_NAMES:
                if cc in label:
                    out['tender_credit_card'] += amt
                    matched = True
                    break
            if matched:
                continue
            if 'CASH' in label:
                out['tender_cash'] += amt
                continue
            for d in DIGITAL_NAMES:
                if d in label:
                    out['tender_digital'] += amt
                    matched = True
                    break
            if matched:
                continue
            for p in PLATFORM_NAMES:
                if p in label:
                    out['tender_platform'] += amt
                    matched = True
                    break
            if matched:
                continue
            for v in VOUCHER_NAMES:
                if v in label:
                    out['tender_voucher'] += amt
                    matched = True
                    break
            if matched:
                continue
            out['tender_other'] += amt

    return out


def parse_menu_items(fpath):
    """Parse menuItemReport and return DI/TA sales by category."""
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    result = {'DI': {}, 'TA': {}}
    current_section = None  # 'DI' or 'TA'
    for r in ws.iter_rows(values_only=True):
        if r[0] == 'Dine In':
            current_section = 'DI'
            continue
        if r[0] == 'Take Away':
            current_section = 'TA'
            continue
        if not current_section:
            continue
        # Category row: col A has name, cols B,C are None, col D has qty
        # Item row: col A is None, col B is Code Id, col C has Item Name
        if r[0] and r[1] is None and r[2] is None and isinstance(r[0], str) and r[0] not in ('Category', 'Grand Total'):
            cat = r[0].strip()
            net_sales = num(r[7]) if len(r) > 7 else 0
            if cat and cat != 'Grand Total' and net_sales > 0:
                result[current_section][cat] = result[current_section].get(cat, 0) + net_sales
    return result


def day_of_week(date_str):
    import datetime
    d = datetime.datetime.strptime(date_str, '%Y-%m-%d')
    return d.strftime('%a')


def main():
    all_data = defaultdict(dict)  # {outlet: {date: metrics}}
    menu_data = defaultdict(dict)  # {outlet: {date: {DI:{}, TA:{}}}}
    files = sorted(os.listdir(KEK_DIR))
    for f in files:
        if not f.endswith('.xlsx'):
            continue
        outlet = identify_outlet(f)
        if not outlet:
            continue
        m = re.search(r'(\d{4}-\d{2}-\d{2})', f)
        if not m:
            continue
        date = m.group(1)
        fpath = os.path.join(KEK_DIR, f)
        if 'dailyReport' in f:
            try:
                data = parse_daily(fpath)
                all_data[outlet][date] = data
                print(f"Parsed daily: {outlet} {date}: net={data['net_sales']:.2f}")
            except Exception as e:
                print(f"ERROR parsing {f}: {e}")
        elif 'menuItemReport' in f:
            try:
                data = parse_menu_items(fpath)
                menu_data[outlet][date] = data
            except Exception as e:
                print(f"ERROR parsing menu {f}: {e}")

    # Save to JSON for inspection
    with open('/sessions/youthful-peaceful-faraday/kek_parsed.json', 'w') as fh:
        json.dump({'daily': all_data, 'menu': menu_data}, fh, default=str, indent=2)
    print(f"\nParsed {sum(len(v) for v in all_data.values())} daily reports across {len(all_data)} outlets")
    return all_data, menu_data


if __name__ == '__main__':
    main()
