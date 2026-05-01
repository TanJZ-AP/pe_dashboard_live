"""Build per-view payment breakdown with voucher detail.

Output: DATA.payment_by_view[view] = {
    cash, credit_card, digital, platform, other, voucher_total, vouchers: {name: amt}
}

Drives the Payment Mix Reconciliation table on the dashboard.
Run after pull_latest.py to refresh.
"""
import openpyxl, os, re, json, sys

SCRIPTS = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPTS)
from _dates import month_weeks

KEK_DIR = os.environ.get('KEK_DIR', "/sessions/youthful-peaceful-faraday/mnt/Daily POS Reports")
UPLOADS = os.environ.get('KEK_UPLOADS', "/sessions/youthful-peaceful-faraday/mnt/uploads")
OUTLETS = ['KEK Alexandra', 'KEK Tampines', 'KEK Punggol', 'WOKIN Punggol']

_W = month_weeks()
W1_DATES, W2_DATES, W3_DATES, W4_DATES, W5_DATES = _W['w1'], _W['w2'], _W['w3'], _W['w4'], _W['w5']

def num(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def identify(f):
    for k, v in [('Alexandra', 'KEK Alexandra'), ('Punggol SAFRA', 'KEK Punggol'),
                 ('Tampines Safra KEK', 'KEK Tampines'), ('Wok In Burger', 'WOKIN Punggol')]:
        if k in f: return v
    return None

def normalize_digital(name):
    """Canonical digital tender names — collapse casing variants and group Paylah/PayNow."""
    n = str(name or '').upper().strip()
    if n in ('PAYLAH', 'PAYNOW'): return 'PayLah/PayNow'
    if n == 'GRAB PAY': return 'Grab Pay'
    if n == 'CDC': return 'CDC'
    if n == 'FAVEPAY': return 'FavePay'
    return name.strip()

def normalize_platform(name):
    """Canonical platform tender names."""
    n = str(name or '').upper().strip()
    if n == 'GRAB FOOD': return 'Grab Food'
    if n in ('FOOD PANDA', 'FOODPANDA'): return 'FoodPanda'
    if n == 'ODDLE': return 'Oddle'
    if n == 'DELIVEROO': return 'Deliveroo'
    return name.strip()

def classify(name):
    n = str(name or '').upper().strip()
    voucher_keys = ['SAFRA VOUCHER', 'KEK VOUCHER', 'NS LIFE', 'NS LIFE SG',
                    'STAFF MEAL', 'SAFRA BIRTHDAY VOUCHER', 'STB EVENT',
                    'DEPOSIT REDEEM', 'GIFT CARD/VOUCHER SALES', 'BONUS TOP UP AMOUNT',
                    'ENT', 'KEK DRINK', 'SAFRA (INTERNAL)']
    for v in voucher_keys:
        if v in n:
            return ('voucher', name.strip())
    if n in ('CASH',): return ('cash', None)
    if n in ('VISA', 'MASTER', 'AMEX'): return ('credit_card', None)
    if n in ('PAYNOW', 'GRAB PAY', 'PAYLAH', 'FAVEPAY', 'CDC'): return ('digital', normalize_digital(name))
    if n in ('GRAB FOOD', 'FOOD PANDA', 'FOODPANDA', 'ODDLE', 'DELIVEROO'): return ('platform', normalize_platform(name))
    return ('other', None)

detail_by_date = {}
def add_detail(date, outlet, cat, amount, voucher_name=None):
    if amount <= 0: return
    rec = detail_by_date.setdefault(date, {}).setdefault(outlet, {
        'cash': 0.0, 'credit_card': 0.0, 'digital_total': 0.0, 'platform_total': 0.0, 'other': 0.0,
        'voucher_total': 0.0, 'vouchers': {}, 'digitals': {}, 'platforms': {}
    })
    if cat == 'voucher':
        rec['voucher_total'] += amount
        rec['vouchers'][voucher_name] = rec['vouchers'].get(voucher_name, 0) + amount
    elif cat == 'digital':
        rec['digital_total'] += amount
        # voucher_name carries the original tender name when not a voucher
        n = voucher_name or 'Digital'
        rec['digitals'][n] = rec['digitals'].get(n, 0) + amount
    elif cat == 'platform':
        rec['platform_total'] += amount
        n = voucher_name or 'Platform'
        rec['platforms'][n] = rec['platforms'].get(n, 0) + amount
    else:
        rec[cat] += amount

# 1. Per-day individual dailyReports — Tender Type section
for f in sorted(os.listdir(KEK_DIR)):
    if 'dailyReport' not in f: continue
    outlet = identify(f)
    if not outlet: continue
    m = re.search(r'(\d{4}-\d{2}-\d{2})', f)
    if not m: continue
    date = m.group(1)
    try:
        wb = openpyxl.load_workbook(os.path.join(KEK_DIR, f), data_only=True)
        ws = wb.active
        in_tender = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Tender Type':
                in_tender = True; continue
            if in_tender and row[0] and isinstance(row[0], str):
                if row[0].startswith('Major Group') or row[0].startswith('Discount Title'):
                    in_tender = False; continue
                if row[0] == 'TOTAL.': continue
                tname = row[0].split('.')[0].strip()
                amt = num(row[1])
                if amt <= 0: continue
                cat, vname = classify(tname)
                add_detail(date, outlet, cat, amt, vname)
    except: pass

# 2. Weekly summary dailyReports — for outlet-days not covered above
WEEKLY_SUMMARIES = [
    ('KEK Alexandra', f"{UPLOADS}/dailyReport_2026.03.30~2026.04.05_Keng Eng Kee Seafood @ Alexandra Village.xlsx"),
    ('KEK Alexandra', f"{UPLOADS}/dailyReport_2026.04.06~2026.04.12_Keng Eng Kee Seafood @ Alexandra Village.xlsx"),
    ('KEK Tampines',  f"{UPLOADS}/dailyReport_2026.03.30~2026.04.05_KEK2 Pte Ltd .xlsx"),
    ('KEK Tampines',  f"{UPLOADS}/dailyReport_2026.04.06~2026.04.12_KEK2 Pte Ltd .xlsx"),
    ('KEK Punggol',   f"{UPLOADS}/dailyReport_2026.03.30~2026.04.04_Keng Eng Kee Seafood @ Punggol SAFRA.xlsx"),
    ('KEK Punggol',   f"{UPLOADS}/dailyReport_2026.04.06~2026.04.12_Keng Eng Kee Seafood @ Punggol SAFRA.xlsx"),
    ('WOKIN Punggol', f"{UPLOADS}/dailyReport_2026.03.30~2026.04.12_WOKIN.xlsx"),
]
for outlet, fpath in WEEKLY_SUMMARIES:
    if not os.path.exists(fpath): continue
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[4]
    for row in rows[5:]:
        if row[0] == 'Total' or not row[0]: continue
        try:
            d, m_, y = str(row[0]).split('-')
            iso = f"{int(y):04d}-{int(m_):02d}-{int(d):02d}"
        except: continue
        existing = detail_by_date.get(iso, {}).get(outlet, {})
        if any(existing.get(k, 0) > 0 for k in ('cash','credit_card','digital','platform','other','voucher_total')):
            continue  # individual file already populated; skip
        for ci, h in enumerate(headers):
            if ci < 15 or h is None or h == '': continue
            if h in ('Deposit Amount', 'Delivery Fee'): continue
            amt = num(row[ci] if ci < len(row) else 0)
            if amt <= 0: continue
            cat, vname = classify(h)
            add_detail(iso, outlet, cat, amt, vname)

def aggregate_view(dates):
    out = {'cash': 0, 'credit_card': 0, 'digital_total': 0, 'platform_total': 0, 'other': 0, 'voucher_total': 0,
           'vouchers': {}, 'digitals': {}, 'platforms': {}}
    for d in dates:
        for outlet, rec in detail_by_date.get(d, {}).items():
            for k in ('cash', 'credit_card', 'digital_total', 'platform_total', 'other', 'voucher_total'):
                out[k] += rec.get(k, 0)
            for vname, vamt in rec.get('vouchers', {}).items():
                out['vouchers'][vname] = out['vouchers'].get(vname, 0) + vamt
            for dn, da in rec.get('digitals', {}).items():
                out['digitals'][dn] = out['digitals'].get(dn, 0) + da
            for pn, pa in rec.get('platforms', {}).items():
                out['platforms'][pn] = out['platforms'].get(pn, 0) + pa
    return out

all_dates = sorted(detail_by_date.keys())
views = {'mtd': all_dates, 'w1': W1_DATES, 'w2': W2_DATES, 'w3': W3_DATES, 'w4': W4_DATES, 'w5': W5_DATES}

def round_pay(p):
    p2 = {k: round(v, 2) for k, v in p.items() if k not in ('vouchers', 'digitals', 'platforms')}
    p2['vouchers'] = {k: round(v, 2) for k, v in p['vouchers'].items()}
    p2['digitals'] = {k: round(v, 2) for k, v in p['digitals'].items()}
    p2['platforms'] = {k: round(v, 2) for k, v in p['platforms'].items()}
    return p2

payment_by_view = {v: round_pay(aggregate_view(dates)) for v, dates in views.items()}

with open(f'{SCRIPTS}/kek_dashboard_data.json') as f:
    dash = json.load(f)
dash['payment_by_view'] = payment_by_view
with open(f'{SCRIPTS}/kek_dashboard_data.json', 'w') as f:
    json.dump(dash, f, indent=2, ensure_ascii=False)

print(f"Payment detail aggregated across {len(views)} views from {len(all_dates)} dates")
for v, p in payment_by_view.items():
    total = p['cash']+p['credit_card']+p['digital_total']+p['platform_total']+p['other']+p['voucher_total']
    print(f"  {v}: total=${total:,.2f} | vouchers: {len(p['vouchers'])} | digitals: {len(p['digitals'])} | platforms: {len(p['platforms'])}")
